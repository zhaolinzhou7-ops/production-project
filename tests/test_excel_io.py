"""Excel 模板生成与导入测试。"""
from __future__ import annotations

import io

from openpyxl import load_workbook

from backend import crud, excel_io
from backend.excel_io import (
    SHEET_MACHINES,
    SHEET_ORDERS,
    SHEET_ROUTES,
    SHEET_SETUP,
)


def test_template_has_all_sheets():
    content = excel_io.build_template()
    wb = load_workbook(io.BytesIO(content))
    assert set(wb.sheetnames) == {
        SHEET_MACHINES, SHEET_SETUP, SHEET_ORDERS, SHEET_ROUTES,
    }


def test_import_template_roundtrip(session):
    """模板自带的示例数据应能直接导入成功。"""
    report = excel_io.import_workbook(session, excel_io.build_template())
    assert report.ok, report.errors
    assert report.machines == 2
    assert report.orders == 1
    assert report.operations == 2
    session.commit()

    order = crud.get_order(session, "J1")
    assert order.operations[0].machines == {"M1": 40, "M2": 50}
    m1 = crud.get_machine(session, "M1")
    assert m1.setup_times == {"A": {"B": 30}, "B": {"A": 20}}


def test_import_reports_row_errors(session):
    """错误行应逐条报告且整体不入库。"""
    wb = load_workbook(io.BytesIO(excel_io.build_template()))
    ws = wb[SHEET_ROUTES]
    ws.append(["J1", 2, "钻孔", "A", "M_NOT_EXIST", 25])   # 机台不存在
    ws.append(["J_NOT_EXIST", 0, "车", "A", "M1", 10])     # 订单不存在
    ws2 = wb[SHEET_ORDERS]
    ws2.append(["J2", "无效交期订单", "not-a-date", 1, "", "normal"])

    buf = io.BytesIO()
    wb.save(buf)
    report = excel_io.import_workbook(session, buf.getvalue())

    assert not report.ok
    messages = "\n".join(e.message for e in report.errors)
    assert "M_NOT_EXIST" in messages
    assert "J_NOT_EXIST" in messages
    assert "交期" in messages
    # 全有或全无: 库应保持为空
    session.rollback()
    assert crud.list_machines(session) == []


def test_import_replace_clears_previous(session):
    excel_io.import_workbook(session, excel_io.build_template())
    session.commit()

    # 改订单 ID 后 replace 导入, 旧订单应消失
    wb = load_workbook(io.BytesIO(excel_io.build_template()))
    ws = wb[SHEET_ORDERS]
    ws.cell(row=2, column=1, value="J99")
    ws_r = wb[SHEET_ROUTES]
    for row in range(2, 5):
        ws_r.cell(row=row, column=1, value="J99")
    buf = io.BytesIO()
    wb.save(buf)

    report = excel_io.import_workbook(session, buf.getvalue(), mode="replace")
    assert report.ok, report.errors
    session.commit()
    assert crud.get_order(session, "J1") is None
    assert crud.get_order(session, "J99") is not None


def test_import_append_upserts(session):
    excel_io.import_workbook(session, excel_io.build_template())
    session.commit()

    wb = load_workbook(io.BytesIO(excel_io.build_template()))
    wb[SHEET_ORDERS].cell(row=2, column=2, value="轴承座-改名")
    buf = io.BytesIO()
    wb.save(buf)

    report = excel_io.import_workbook(session, buf.getvalue(), mode="append")
    assert report.ok, report.errors
    session.commit()
    assert crud.get_order(session, "J1").name == "轴承座-改名"
