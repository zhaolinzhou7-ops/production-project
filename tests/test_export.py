"""方案 Excel 导出测试 (导出后回读断言)。"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import load_workbook

from backend import excel_io
from backend.scenario import save_scenario

from test_reports import _fake_result


def test_export_scenario_roundtrip(session):
    summary = save_scenario(
        session, name="导出测试", kind="baseline", result=_fake_result(),
        schedule_start=datetime(2026, 7, 6, 8, 0),
    )
    session.commit()

    content = excel_io.export_scenario(session, summary.id)
    assert content is not None
    wb = load_workbook(io.BytesIO(content))
    assert wb.sheetnames == ["排产明细", "订单交付", "机台利用"]

    detail = list(wb["排产明细"].iter_rows(values_only=True))
    assert len(detail) == 1 + 3  # 表头 + 3 道工序
    # 外协工序机台列显示 "外协"
    assert any(row[4] == "外协" for row in detail[1:])
    # 绝对时间: 08:00 开工
    assert detail[1][6] == "2026-07-06 08:00"

    delivery = list(wb["订单交付"].iter_rows(values_only=True))
    assert len(delivery) == 1 + 2
    risks = {row[0]: row[6] for row in delivery[1:]}
    assert risks["J1"] == "拖期"
    assert risks["J2"] == "安全"

    util = list(wb["机台利用"].iter_rows(values_only=True))
    assert util[1][0] == "M1"
    assert util[1][1] == 200


def test_export_missing_scenario(session):
    assert excel_io.export_scenario(session, 999) is None
