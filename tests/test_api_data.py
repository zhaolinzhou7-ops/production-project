"""数据管理 API 集成测试 (TestClient + 内存 SQLite)。"""
from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import sessionmaker

from backend.db import create_db_engine, get_session
from backend.main import app

MACHINE = {
    "id": "M1", "name": "CNC 车床",
    "setup_times": {"A": {"B": 30}, "B": {"A": 20}},
}
ORDER = {
    "id": "J1", "name": "轴承座", "due_date": "2026-07-10T17:00:00", "priority": 3,
    "operations": [
        {"seq": 0, "name": "粗车", "family": "A", "machines": {"M1": 40}},
        {"seq": 1, "name": "精铣", "family": "A", "machines": {"M1": 35}},
    ],
}


@pytest.fixture()
def client():
    engine = create_db_engine("sqlite:///:memory:")
    factory = sessionmaker(bind=engine, expire_on_commit=False)

    def override():
        s = factory()
        try:
            yield s
            s.commit()
        except Exception:
            s.rollback()
            raise
        finally:
            s.close()

    app.dependency_overrides[get_session] = override
    yield TestClient(app)
    app.dependency_overrides.clear()
    engine.dispose()


def test_machine_crud_flow(client):
    assert client.post("/api/machines", json=MACHINE).status_code == 200
    # 重复创建 409
    assert client.post("/api/machines", json=MACHINE).status_code == 409

    res = client.get("/api/machines")
    assert [m["id"] for m in res.json()] == ["M1"]

    res = client.put(
        "/api/machines/M1/setup-times", json={"A": {"B": 45}}
    )
    assert res.json()["setup_times"] == {"A": {"B": 45}}

    assert client.delete("/api/machines/M1").status_code == 200
    assert client.get("/api/machines").json() == []


def test_order_crud_flow(client):
    client.post("/api/machines", json=MACHINE)
    res = client.post("/api/orders", json=ORDER)
    assert res.status_code == 200, res.text
    assert res.json()["operations"][0]["id"] == "J1-0"

    # 引用不存在机台 -> 422
    bad = {**ORDER, "id": "J2",
           "operations": [{"seq": 0, "name": "x", "family": "A",
                           "machines": {"MX": 10}}]}
    assert client.post("/api/orders", json=bad).status_code == 422

    # 机台被引用时删除 -> 409
    assert client.delete("/api/machines/M1").status_code == 409

    assert client.delete("/api/orders/J1").status_code == 200
    assert client.get("/api/orders/J1").status_code == 404


def test_template_download_and_import(client):
    res = client.get("/api/import/template")
    assert res.status_code == 200
    wb = load_workbook(io.BytesIO(res.content))
    assert "订单" in wb.sheetnames

    res = client.post(
        "/api/import/excel?mode=replace",
        files={"file": ("t.xlsx", res.content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    report = res.json()
    assert report["ok"], report["errors"]
    assert report["machines"] == 2

    assert len(client.get("/api/orders").json()) == 1


def test_schedule_request_from_db_and_solve(client):
    """端到端: 导入模板 -> 组装请求 -> 排产成功。"""
    template = client.get("/api/import/template").content
    client.post(
        "/api/import/excel?mode=replace",
        files={"file": ("t.xlsx", template, "application/octet-stream")},
    )

    res = client.get(
        "/api/data/schedule-request",
        params={"schedule_start": "2026-07-01T08:00:00", "time_limit_seconds": 5},
    )
    assert res.status_code == 200, res.text
    request = res.json()
    assert len(request["machines"]) == 2

    solved = client.post("/api/schedule", json=request)
    assert solved.status_code == 200
    body = solved.json()
    assert body["status"] in ("OPTIMAL", "FEASIBLE")
    assert len(body["operations"]) == 2


def test_schedule_request_empty_db_422(client):
    assert client.get("/api/data/schedule-request").status_code == 422
