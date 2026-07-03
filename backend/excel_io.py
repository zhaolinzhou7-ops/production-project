"""Excel 模板生成与导入 (openpyxl)。

模板含 4 个 sheet: 机台 / 换型矩阵 / 订单 / 工艺路线。
- 工艺路线一行 = (订单, 工序) 在一台机台上的可选加工, 同工序多台机台写多行;
- 导入为全量校验、全有或全无: 任一行出错则整体不入库, 返回逐行错误报告。
"""
from __future__ import annotations

import io
from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from . import crud
from .schemas import (
    CalendarDTO,
    CalendarExceptionDTO,
    ImportIssue,
    ImportReport,
    MachineDTO,
    OperationDTO,
    OrderDTO,
    ResourceDTO,
    ShiftRuleDTO,
)

SHEET_MACHINES = "机台"
SHEET_SETUP = "换型矩阵"
SHEET_ORDERS = "订单"
SHEET_ROUTES = "工艺路线"
SHEET_CALENDARS = "日历"
SHEET_RESOURCES = "资源"

MACHINE_HEADERS = ["机台ID", "机台名称", "是否启用(是/否)", "日历ID(可空)"]
SETUP_HEADERS = ["机台ID", "原产品族", "新产品族", "换型时间(分钟)"]
ORDER_HEADERS = [
    "订单ID", "订单名称", "交期(YYYY-MM-DD HH:MM)", "优先级(>=1)",
    "最早开工(可空)", "状态(normal/rush)",
]
ROUTE_HEADERS = [
    "订单ID", "工序号(0起)", "工序名称", "产品族", "机台ID", "加工时长(分钟)",
    "外协周期(分钟,填了即外协)", "资源ID(可空)", "资源数量",
]
CALENDAR_HEADERS = [
    "日历ID", "日历名称", "星期(1-7,或日期YYYY-MM-DD)", "班次开始(HH:MM)",
    "班次结束(HH:MM)", "整天休(是/否,仅日期行)",
]
RESOURCE_HEADERS = ["资源ID", "资源名称", "同时可用数量"]

DATE_FORMATS = ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d")


def build_template() -> bytes:
    """生成带示例数据的导入模板。"""
    wb = Workbook()

    ws = wb.active
    ws.title = SHEET_MACHINES
    ws.append(MACHINE_HEADERS)
    ws.append(["M1", "CNC 车床", "是", "C1"])
    ws.append(["M2", "加工中心", "是", ""])

    ws = wb.create_sheet(SHEET_SETUP)
    ws.append(SETUP_HEADERS)
    ws.append(["M1", "A", "B", 30])
    ws.append(["M1", "B", "A", 20])

    ws = wb.create_sheet(SHEET_ORDERS)
    ws.append(ORDER_HEADERS)
    ws.append(["J1", "轴承座", "2026-07-10 17:00", 3, "", "normal"])

    ws = wb.create_sheet(SHEET_ROUTES)
    ws.append(ROUTE_HEADERS)
    ws.append(["J1", 0, "粗车", "A", "M1", 40, "", "", ""])
    ws.append(["J1", 0, "粗车", "A", "M2", 50, "", "", ""])
    ws.append(["J1", 1, "热处理", "A", "", "", 240, "", ""])
    ws.append(["J1", 2, "精铣", "A", "M2", 35, "", "F1", 1])

    ws = wb.create_sheet(SHEET_CALENDARS)
    ws.append(CALENDAR_HEADERS)
    for wd in (1, 2, 3, 4, 5):
        ws.append(["C1", "两班制", wd, "08:00", "22:00", ""])
    ws.append(["C1", "两班制", "2026-10-01", "", "", "是"])

    ws = wb.create_sheet(SHEET_RESOURCES)
    ws.append(RESOURCE_HEADERS)
    ws.append(["F1", "专用夹具", 1])

    for name in wb.sheetnames:
        for col in wb[name].columns:
            letter = col[0].column_letter
            wb[name].column_dimensions[letter].width = max(
                14, max(len(str(c.value or "")) for c in col) + 4
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---- 方案结果导出 -----------------------------------------------------------

def export_scenario(session: Session, scenario_id: int) -> bytes | None:
    """把方案导出为 Excel: 排产明细 / 订单交付 / 机台利用 三个 sheet。"""
    from . import reports, scenario as scenario_mod
    from .models import OUTSOURCE_MACHINE_ID

    detail = scenario_mod.get_scenario(session, scenario_id)
    if detail is None:
        return None
    start = detail.schedule_start

    def dt(minutes: int) -> str:
        return (start + timedelta(minutes=minutes)).strftime("%Y-%m-%d %H:%M")

    wb = Workbook()
    ws = wb.active
    ws.title = "排产明细"
    ws.append(["订单", "订单名称", "工序", "工序名称", "机台", "产品族",
               "开工", "完工", "加工(分钟)", "换型(分钟)", "冻结"])
    for op in sorted(detail.operations, key=lambda x: (x["machine_id"], x["start"])):
        ws.append([
            op["order_id"], op["order_name"], op["operation_id"],
            op["operation_name"],
            "外协" if op["machine_id"] == OUTSOURCE_MACHINE_ID else op["machine_id"],
            op["family"], dt(op["start"]), dt(op["end"]),
            op["duration"], op["setup"], "是" if op.get("frozen") else "",
        ])

    ws = wb.create_sheet("订单交付")
    ws.append(["订单", "名称", "交期", "预计完工", "拖期(分钟)", "富余(分钟)", "风险"])
    risk = reports.delivery_risk(session, scenario_id)
    risk_names = {"red": "拖期", "yellow": "紧张", "green": "安全"}
    for o in risk.orders:
        ws.append([
            o.order_id, o.order_name,
            o.due.strftime("%Y-%m-%d %H:%M"),
            o.completion.strftime("%Y-%m-%d %H:%M"),
            o.tardiness_min, o.slack_min, risk_names[o.risk],
        ])

    ws = wb.create_sheet("机台利用")
    ws.append(["机台", "占用(分钟)", "利用率"])
    for m in reports.kpi(session, scenario_id).machines:
        ws.append([m.machine_id, m.busy_min, m.utilization])

    for name in wb.sheetnames:
        for col in wb[name].columns:
            letter = col[0].column_letter
            wb[name].column_dimensions[letter].width = max(
                12, max(len(str(c.value or "")) for c in col) + 4
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---- 导入解析 ---------------------------------------------------------------

def _cell_str(v) -> str:
    return str(v).strip() if v is not None else ""


def _parse_int(v, field: str) -> int:
    try:
        n = int(float(str(v)))
    except (TypeError, ValueError):
        raise ValueError(f"{field} 必须为整数, 实际为 {v!r}")
    return n


def _parse_dt(v, field: str) -> datetime:
    if isinstance(v, datetime):
        return v
    s = _cell_str(v)
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f"{field} 无法解析为日期时间: {v!r} (格式如 2026-07-10 17:00)")


def _data_rows(ws: Worksheet):
    """跳过表头, 返回 (行号, 值元组), 忽略整行为空的行。"""
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None or _cell_str(c) == "" for c in row):
            continue
        yield idx, row


def import_workbook(session: Session, content: bytes, mode: str = "replace") -> ImportReport:
    """导入 Excel。mode=replace 清库重建, append 按 ID 覆盖/追加。"""
    if mode not in ("replace", "append"):
        raise ValueError(f"未知导入模式: {mode}")

    errors: list[ImportIssue] = []

    def err(sheet: str, row: int, message: str) -> None:
        errors.append(ImportIssue(sheet=sheet, row=row, message=message))

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:  # noqa: BLE001 - 文件级错误统一报告
        return ImportReport(ok=False, mode=mode, errors=[
            ImportIssue(sheet="-", row=0, message=f"无法读取 Excel 文件: {e}")
        ])

    for name in (SHEET_MACHINES, SHEET_ORDERS, SHEET_ROUTES):
        if name not in wb.sheetnames:
            err(name, 0, f"缺少工作表「{name}」")
    if errors:
        return ImportReport(ok=False, mode=mode, errors=errors)

    # -- 日历 (可选 sheet): 星期行=周规则, 日期行=例外 --
    calendars: dict[str, CalendarDTO] = {}
    if SHEET_CALENDARS in wb.sheetnames:
        for r, row in _data_rows(wb[SHEET_CALENDARS]):
            cid = _cell_str(row[0])
            cname = _cell_str(row[1] if len(row) > 1 else "")
            day_s = _cell_str(row[2] if len(row) > 2 else "")
            start_s = _cell_str(row[3] if len(row) > 3 else "")
            end_s = _cell_str(row[4] if len(row) > 4 else "")
            off_s = _cell_str(row[5] if len(row) > 5 else "")
            if not cid or not day_s:
                err(SHEET_CALENDARS, r, "日历ID 与 星期/日期 不能为空")
                continue
            cal = calendars.setdefault(cid, CalendarDTO(id=cid, name=cname))
            try:
                if "-" in day_s or isinstance(row[2], datetime):
                    day = (
                        row[2].strftime("%Y-%m-%d")
                        if isinstance(row[2], datetime) else day_s
                    )
                    off = off_s in ("是", "yes", "true", "1")
                    cal.exceptions.append(CalendarExceptionDTO(
                        date=day, available=not off,
                        start=start_s or None, end=end_s or None,
                    ))
                else:
                    wd = int(float(day_s))
                    if not 1 <= wd <= 7:
                        raise ValueError(f"星期必须为 1-7, 实际为 {day_s}")
                    if not start_s or not end_s:
                        raise ValueError("周规则行必须填写班次开始/结束")
                    cal.rules.append(ShiftRuleDTO(
                        weekday=wd - 1, start=start_s, end=end_s,
                    ))
            except ValueError as e:
                err(SHEET_CALENDARS, r, str(e))

    # -- 资源 (可选 sheet) --
    resources: dict[str, ResourceDTO] = {}
    if SHEET_RESOURCES in wb.sheetnames:
        for r, row in _data_rows(wb[SHEET_RESOURCES]):
            rid = _cell_str(row[0])
            if not rid:
                err(SHEET_RESOURCES, r, "资源ID 不能为空")
                continue
            try:
                cap = _parse_int(row[2], "同时可用数量") if len(row) > 2 and _cell_str(row[2]) else 1
                if cap < 1:
                    raise ValueError("同时可用数量必须 >= 1")
            except ValueError as e:
                err(SHEET_RESOURCES, r, str(e))
                continue
            resources[rid] = ResourceDTO(
                id=rid, name=_cell_str(row[1] if len(row) > 1 else ""), capacity=cap,
            )

    # -- 机台 --
    machines: dict[str, MachineDTO] = {}
    for r, row in _data_rows(wb[SHEET_MACHINES]):
        mid, name = _cell_str(row[0]), _cell_str(row[1] if len(row) > 1 else "")
        active_s = _cell_str(row[2] if len(row) > 2 else "是") or "是"
        cal_id = _cell_str(row[3] if len(row) > 3 else "")
        if not mid or not name:
            err(SHEET_MACHINES, r, "机台ID 与 机台名称 不能为空")
            continue
        if mid in machines:
            err(SHEET_MACHINES, r, f"机台ID 重复: {mid}")
            continue
        if cal_id and cal_id not in calendars:
            err(SHEET_MACHINES, r, f"日历ID 不存在: {cal_id}")
            continue
        machines[mid] = MachineDTO(
            id=mid, name=name, active=active_s not in ("否", "no", "false", "0"),
            calendar_id=cal_id or None,
        )

    # -- 换型矩阵 --
    if SHEET_SETUP in wb.sheetnames:
        for r, row in _data_rows(wb[SHEET_SETUP]):
            mid = _cell_str(row[0])
            fam_from = _cell_str(row[1] if len(row) > 1 else "")
            fam_to = _cell_str(row[2] if len(row) > 2 else "")
            if mid not in machines:
                err(SHEET_SETUP, r, f"机台ID 不存在: {mid}")
                continue
            if not fam_from or not fam_to:
                err(SHEET_SETUP, r, "原产品族/新产品族不能为空")
                continue
            try:
                minutes = _parse_int(row[3] if len(row) > 3 else None, "换型时间")
                if minutes < 0:
                    raise ValueError("换型时间不能为负")
            except ValueError as e:
                err(SHEET_SETUP, r, str(e))
                continue
            machines[mid].setup_times.setdefault(fam_from, {})[fam_to] = minutes

    # -- 订单 --
    order_meta: dict[str, dict] = {}
    for r, row in _data_rows(wb[SHEET_ORDERS]):
        oid, name = _cell_str(row[0]), _cell_str(row[1] if len(row) > 1 else "")
        if not oid or not name:
            err(SHEET_ORDERS, r, "订单ID 与 订单名称 不能为空")
            continue
        if oid in order_meta:
            err(SHEET_ORDERS, r, f"订单ID 重复: {oid}")
            continue
        try:
            due = _parse_dt(row[2] if len(row) > 2 else None, "交期")
            priority = _parse_int(row[3], "优先级") if len(row) > 3 and _cell_str(row[3]) else 1
            release_v = row[4] if len(row) > 4 else None
            release = _parse_dt(release_v, "最早开工") if _cell_str(release_v) else None
            status = _cell_str(row[5] if len(row) > 5 else "") or "normal"
            if priority < 1:
                raise ValueError("优先级必须 >= 1")
            if status not in ("normal", "rush"):
                raise ValueError(f"状态只能为 normal/rush, 实际为 {status!r}")
        except ValueError as e:
            err(SHEET_ORDERS, r, str(e))
            continue
        order_meta[oid] = dict(
            name=name, due_date=due, priority=priority,
            release_time=release, status=status, ops={},
        )

    # -- 工艺路线 --
    for r, row in _data_rows(wb[SHEET_ROUTES]):
        oid = _cell_str(row[0])
        if oid not in order_meta:
            err(SHEET_ROUTES, r, f"订单ID 不存在: {oid}")
            continue
        try:
            seq = _parse_int(row[1] if len(row) > 1 else None, "工序号")
            name = _cell_str(row[2] if len(row) > 2 else "")
            family = _cell_str(row[3] if len(row) > 3 else "")
            mid = _cell_str(row[4] if len(row) > 4 else "")
            lead_s = _cell_str(row[6] if len(row) > 6 else "")
            res_id = _cell_str(row[7] if len(row) > 7 else "") or None
            res_qty = (
                _parse_int(row[8], "资源数量")
                if len(row) > 8 and _cell_str(row[8]) else 1
            )
            if not name or not family:
                raise ValueError("工序名称/产品族不能为空")
            if res_id and res_id not in resources:
                raise ValueError(f"资源ID 不存在: {res_id}")
            if lead_s:  # 外协工序: 忽略机台/时长
                lead = _parse_int(lead_s, "外协周期")
                if lead <= 0:
                    raise ValueError("外协周期必须为正")
                mid, dur = None, None
            else:
                dur = _parse_int(row[5] if len(row) > 5 else None, "加工时长")
                lead = None
                if mid not in machines:
                    raise ValueError(f"机台ID 不存在: {mid}")
                if dur <= 0:
                    raise ValueError("加工时长必须为正")
        except ValueError as e:
            err(SHEET_ROUTES, r, str(e))
            continue
        ops = order_meta[oid]["ops"]
        op = ops.setdefault(seq, dict(
            name=name, family=family, machines={},
            is_outsourced=lead is not None, outsource_lead_min=lead,
            resource_id=res_id, resource_qty=res_qty,
        ))
        if op["name"] != name or op["family"] != family:
            err(SHEET_ROUTES, r, f"订单 {oid} 工序 {seq} 的名称/产品族与前面行不一致")
            continue
        if lead is not None:
            if op["machines"]:
                err(SHEET_ROUTES, r, f"订单 {oid} 工序 {seq} 既有机台行又有外协行")
            continue
        if op["is_outsourced"]:
            err(SHEET_ROUTES, r, f"订单 {oid} 工序 {seq} 既有外协行又有机台行")
            continue
        if mid in op["machines"]:
            err(SHEET_ROUTES, r, f"订单 {oid} 工序 {seq} 重复指定机台 {mid}")
            continue
        op["machines"][mid] = dur

    # -- 组装订单 DTO (借用 Pydantic 校验) --
    orders: list[OrderDTO] = []
    for oid, meta in order_meta.items():
        if not meta["ops"]:
            err(SHEET_ORDERS, 0, f"订单 {oid} 在工艺路线中没有任何工序")
            continue
        try:
            orders.append(OrderDTO(
                id=oid, name=meta["name"], due_date=meta["due_date"],
                priority=meta["priority"], release_time=meta["release_time"],
                status=meta["status"],
                operations=[
                    OperationDTO(seq=seq, **op) for seq, op in sorted(meta["ops"].items())
                ],
            ))
        except ValueError as e:
            err(SHEET_ROUTES, 0, f"订单 {oid}: {e}")

    if errors:
        return ImportReport(ok=False, mode=mode, errors=errors)

    # -- 入库 (全有或全无) --
    if mode == "replace":
        crud.clear_all(session)
    for cal in calendars.values():
        crud.upsert_calendar(session, cal)
    for res in resources.values():
        crud.upsert_resource(session, res)
    for m in machines.values():
        crud.upsert_machine(session, m)
    total_ops = 0
    for o in orders:
        crud.upsert_order(session, o)
        total_ops += len(o.operations)

    return ImportReport(
        ok=True, mode=mode,
        machines=len(machines), orders=len(orders), operations=total_ops,
    )
